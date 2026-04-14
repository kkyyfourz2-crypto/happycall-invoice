from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
import base64
import json
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# ===== 월말 수요처 매핑 =====
WOLMAL_MAP = {
    '케이엔유': {'aliases': ['knu','케이엔유','안순례'], 'addr': ['영동대로86길','동남유화']},
    '갓텐코리아': {'aliases': ['갓텐','갓덴','갓덴스시','갓텐스시'], 'addr': ['역삼로7길','평원빌딩']},
    '경성화원': {'aliases': ['경성화원','경성'], 'addr': ['퇴계로217','진양꽃상가']},
    '차병원 파라메딕': {'aliases': ['차병원','파라메딕'], 'addr': ['서초대로70길','동탄원천로']},
    '까사데이피오리': {'aliases': ['까사데이피오리','까사','까사데'], 'addr': ['논현로841']},
    '더부케(꽃뜰)': {'aliases': ['꽃뜰','어반플라워','어반','더부케'], 'addr': ['동산로8길']},
    '동방미인': {'aliases': ['동방미인','동방'], 'addr': ['봉은사로18길']},
    '더라임플라워': {'aliases': ['더라임플라워','라임플라워','라임','더라임'], 'addr': ['강남대로10길']},
    '리첸플라워': {'aliases': ['리첸플라워','리첸꽃집','리첸'], 'addr': ['양재동232','지하84호']},
    '메디넥스': {'aliases': ['메디넥스'], 'addr': ['노해로69길','세정빌딩']},
    '우리은행(삼성타운점)': {'aliases': ['우리은행','우리은행기업경영','우리은행기업경영지원팀'], 'addr': ['서초대로74길11','삼성전자']},
    '해바라기': {'aliases': ['해바라기'], 'addr': ['서초대로46길67']},
    '플라워센터': {'aliases': ['플라워센터'], 'addr': ['신반포로162','르본시티','39호']},
    '미니쉬치과': {'aliases': ['미니쉬치과','미니쉬'], 'addr': ['언주로728','미니쉬빌딩']},
    '이안동물의학센터': {'aliases': ['이안동물의학센터','이안'], 'addr': ['선릉로806','킹콩빌딩']},
    '퓨얼리영': {'aliases': ['퓨얼리영','퓨얼리'], 'addr': ['한강대로401']},
    '플러스플라워': {'aliases': ['플러스플라워','플러스','엘린플라워'], 'addr': ['남부순환로2716']},
    '마그넷서비스서초지점': {'aliases': ['마그넷서비스','마그넷'], 'addr': []},
    '꽃눈': {'aliases': ['꽃눈'], 'addr': ['오금로11길']},
    '화란플라워': {'aliases': ['화란플라워','화란'], 'addr': ['강남대로27','지하29호']},
    '플로라빌': {'aliases': ['플로라빌'], 'addr': ['세종대로9길42']},
    '신한은행': {'aliases': ['신한은행','신한'], 'addr': ['서초대로74길4','삼성생명']},
    '애플대리점': {'aliases': ['애플대리점','애플'], 'addr': ['신촌로109']},
    '이지메디컴': {'aliases': ['이지메디컴'], 'addr': []},
    '마포르플로스': {'aliases': ['마포르플로스','르플로','마포르'], 'addr': ['삼개로23']},
    '핀플라워': {'aliases': ['핀플라워'], 'addr': ['송이로28길']},
}

# ===== 배송원 전화번호 매핑 =====
PHONE_TO_DRIVER = {
    '01067008058': '김경숙', '01052834669': '김주화',
    '01020577244': '김현',   '01039243430': '김호년',
    '01044180066': '남명현', '01045803928': '남연순',
    '01053938313': '노상운', '01084231532': '문병실',
    '01032492589': '박규호', '01065651723': '박용욱',
    '01020716424': '안영권', '01062516111': '오승우',
    '01091447181': '오인성', '01093876828': '원명보',
    '01086072043': '윤기상', '01077366443': '윤승소',
    '01023771434': '이원근', '01032012413': '이종화',
    '01032324514': '이진영', '01087174575': '이창만',
    '01056552901': '전인석', '01034610228': '전창식',
    '01027201234': '정준서', '01062067732': '최성재',
    '01087403298': '최용주', '01087634482': '허기석',
}

def match_driver_by_phone(phone_str):
    if not phone_str:
        return None
    clean = ''.join(filter(str.isdigit, str(phone_str)))
    return PHONE_TO_DRIVER.get(clean)

def match_vendor(name_str, addr_str=''):
    n = (name_str or '').replace(' ', '').lower()
    a = (addr_str or '').replace(' ', '').lower()
    for std, info in WOLMAL_MAP.items():
        for alias in info['aliases']:
            if alias.replace(' ', '').lower() in n and alias:
                return {'name': std, 'confident': True}
        for key in info['addr']:
            if key.replace(' ', '').lower() in a and key:
                return {'name': std, 'confident': True}
    return {'name': name_str or '', 'confident': False}

def normalize_payment(raw):
    r = (raw or '').replace(' ', '').lower()
    if any(x in r for x in ['월말', '月末', '월말결제', '원말', '웜말', '윌말']):
        return '월말'
    if any(x in r for x in ['현금', '선불', '착불', '후불']):
        return '현금'
    if any(x in r for x in ['계좌', '계좌이체', '이체', '무통장']):
        return '계좌'
    return None

# ===== API 라우트 =====

@app.route('/')
def index():
    return send_file('static/index.html')

@app.route('/api/analyze', methods=['POST'])
def analyze():
    data = request.json
    b64 = data.get('image')
    forced_payment = data.get('payment_type')
    anthropic_key = data.get('anthropic_key')
    vision_key = data.get('vision_key')

    if not b64 or not anthropic_key:
        return jsonify({'error': 'API 키와 이미지가 필요합니다'}), 400

    extracted_text = ''

    # 1. Google Vision으로 텍스트 추출
    if vision_key:
        try:
            vision_url = f'https://vision.googleapis.com/v1/images:annotate?key={vision_key}'
            vision_resp = requests.post(vision_url, json={
                'requests': [{
                    'image': {'content': b64},
                    'features': [{'type': 'DOCUMENT_TEXT_DETECTION', 'maxResults': 1}]
                }]
            }, timeout=30)
            vision_data = vision_resp.json()
            extracted_text = vision_data.get('responses', [{}])[0].get('fullTextAnnotation', {}).get('text', '')
        except Exception as e:
            print(f'Vision API 오류: {e}')

    vendor_list = '\n'.join([f"{k}: {', '.join(v['aliases'])}" for k, v in WOLMAL_MAP.items()])
    vision_context = f'\n\n[Google Vision 추출 텍스트]\n{extracted_text}\n' if extracted_text else ''

    prompt = f"""이것은 "해피콜 지하철택배" 배송 전표입니다.{vision_context}

[전표 구조]
- 상단: 날짜 (20○○년 ○월 ○일)
- 좌측: 출발지 상호(주소), 도착지 상호(주소)
- 우측 "기 타( )" 칸: 결제방식이 손글씨로 적혀있음
- 우측 "배송원:" 옆: 배송원 이름(크게 인쇄), 그 아래 "(연락처: 010-XXXX-XXXX)"

[결제방식]
"기 타" 칸의 손글씨: 원말/웜말/윌말/월말/月末 → "월말", 현금/선불/착불/후불 → "현금", 계좌/이체 → "계좌"

[배송원]
"배송원:" 바로 옆 크게 인쇄된 이름. 성+이름 붙여쓰기.
"(연락처: 010-XXXX-XXXX)" 번호를 배송원전화로 추출.
절대로 080-220-9988, 출발지TEL, 도착지TEL을 배송원 번호로 혼동하지 말것.

[금액] 운송요금란 숫자. 1000원 단위. 6000~40000원 사이.

[월말 수요처 표준명]
{vendor_list}

반드시 순수 JSON 배열만 반환:
[{{"날짜":"2026-03-30","결제방식":"월말","수요처":"미니쉬치과","출발지":"미니쉬치과","도착지":"이살리는치과 선릉","가격":8000,"배송원":"남명현","배송원전화":"010-4410-0066","입금자명":"","비고":"","uncertain":false}}]

판독 불가 항목만 uncertain:true. 전표 아니면 [] 반환."""

    # 2. Claude로 데이터 정리
    try:
        claude_resp = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'Content-Type': 'application/json',
                'x-api-key': anthropic_key,
                'anthropic-version': '2023-06-01'
            },
            json={
                'model': 'claude-sonnet-4-20250514',
                'max_tokens': 2000,
                'messages': [{
                    'role': 'user',
                    'content': [
                        {'type': 'image', 'source': {'type': 'base64', 'media_type': 'image/jpeg', 'data': b64}},
                        {'type': 'text', 'text': prompt}
                    ]
                }]
            },
            timeout=60
        )
        claude_data = claude_resp.json()
        if 'error' in claude_data:
            return jsonify({'error': claude_data['error']['message']}), 400

        text = ''.join(c.get('text', '') for c in claude_data.get('content', []))
        clean = text.replace('```json', '').replace('```', '').strip()
        parsed = json.loads(clean)

        # 후처리
        results = []
        for item in parsed:
            # 결제방식 강제 설정
            payment = forced_payment or normalize_payment(item.get('결제방식')) or '월말'
            # 배송원 전화번호로 매칭
            driver = match_driver_by_phone(item.get('배송원전화')) or item.get('배송원', '')
            # 수요처 표준명 매칭
            vm = match_vendor(item.get('수요처', ''), item.get('출발지', ''))

            results.append({
                **item,
                '결제방식': payment,
                '배송원': driver,
                '수요처': vm['name'] if vm['confident'] else item.get('수요처', ''),
                '_uncertain': item.get('uncertain', False) or not vm['confident'],
            })

        return jsonify({'results': results, 'vision_text': extracted_text})

    except json.JSONDecodeError:
        return jsonify({'error': 'JSON 파싱 실패'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export', methods=['POST'])
def export_excel():
    data = request.json
    items = data.get('items', [])
    month = data.get('month', datetime.now().strftime('%Y%m'))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='2563EB', end_color='2563EB')
    red_font = Font(color='FF0000', bold=True)

    def make_sheet(wb, name, headers, rows, uncertain_cols=[]):
        ws = wb.create_sheet(name)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row['data'], 1):
                cell = ws.cell(ri, ci, val)
                if row.get('uncertain') and ci in uncertain_cols:
                    cell.font = red_font
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        return ws

    # 월말 시트
    wolmal = [d for d in items if d.get('결제방식') == '월말']
    make_sheet(wb, '월말',
        ['날짜', '수요처', '출발지', '도착지', '가격', '배송원', '비고'],
        [{'data': [d.get('날짜',''), d.get('수요처',''), d.get('출발지',''), d.get('도착지',''),
                   d.get('가격',0), d.get('배송원',''), d.get('비고','')],
          'uncertain': d.get('_uncertain', False)} for d in wolmal],
        uncertain_cols=[2, 6]
    )

    # 현금 시트
    hyunkum = [d for d in items if d.get('결제방식') == '현금']
    make_sheet(wb, '현금',
        ['날짜', '출발지', '도착지', '금액', '배송원', '비고'],
        [{'data': [d.get('날짜',''), d.get('출발지',''), d.get('도착지',''),
                   d.get('가격',0), d.get('배송원',''), d.get('비고','')],
          'uncertain': d.get('_uncertain', False)} for d in hyunkum],
        uncertain_cols=[5]
    )

    # 계좌 시트
    gyejwa = [d for d in items if d.get('결제방식') == '계좌']
    make_sheet(wb, '계좌',
        ['날짜', '출발지', '도착지', '가격', '배송원', '확인', '입금자명', '비고'],
        [{'data': [d.get('날짜',''), d.get('출발지',''), d.get('도착지',''),
                   d.get('가격',0), d.get('배송원',''), '', d.get('입금자명',''), d.get('비고','')],
          'uncertain': d.get('_uncertain', False)} for d in gyejwa],
        uncertain_cols=[5]
    )

    # 집계 시트
    ws_sum = wb.create_sheet('지하철택배')
    drivers = {}
    for d in items:
        name = d.get('배송원', '')
        if not name:
            continue
        if name not in drivers:
            drivers[name] = {'월말': 0, '월말금액': 0, '현금': 0, '현금금액': 0, '계좌': 0, '계좌금액': 0}
        t = d.get('결제방식', '')
        p = int(d.get('가격', 0) or 0)
        if t == '월말':
            drivers[name]['월말'] += 1
            drivers[name]['월말금액'] += p
        elif t == '현금':
            drivers[name]['현금'] += 1
            drivers[name]['현금금액'] += p
        elif t == '계좌':
            drivers[name]['계좌'] += 1
            drivers[name]['계좌금액'] += p

    headers = ['배송원', '월말(건)', '월말(금액)', '현금(건)', '현금(금액)', '계좌(건)', '계좌(금액)', '합계(건)', '합계(금액)']
    for ci, h in enumerate(headers, 1):
        cell = ws_sum.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for ri, (name, d) in enumerate(sorted(drivers.items()), 2):
        tc = d['월말'] + d['현금'] + d['계좌']
        ta = d['월말금액'] + d['현금금액'] + d['계좌금액']
        row = [name, d['월말'], d['월말금액'], d['현금'], d['현금금액'], d['계좌'], d['계좌금액'], tc, ta]
        for ci, val in enumerate(row, 1):
            ws_sum.cell(ri, ci, val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'해피콜_전표정리_{month}.xlsx'
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
